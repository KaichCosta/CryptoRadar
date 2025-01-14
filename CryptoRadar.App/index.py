import time
from datetime import datetime
import os
import openpyxl
from urllib.parse import quote
import webbrowser
import pyautogui
from tkinter import *
from tkinter import messagebox
import threading
import requests
import sys
# Variáveis globais
thread_manter_sistema = None
start_time = None
parar_sistema_event = threading.Event()

#2 buscar preços bitcoin e ações atuais
precos = {
    "BTC": {"compra": 0, "venda": 0},  # Armazena preço de compra e venda do Bitcoin
    "ETH": {"compra": 0, "venda": 0}   # Armazena preço de compra e venda do Ethereum
}
def atualizar_preco():
    def tarefa():
        global precos, infohora
        while not parar_sistema_event.is_set():
            requisicao = requests.get("https://economia.awesomeapi.com.br/last/BTC-USD,ETH-USD")
            requisicao_dic = requisicao.json()
            precos["BTC"]["compra"] = round(float(requisicao_dic['BTCUSD']['bid']),2)
            precos["BTC"]["venda"] = round(float(requisicao_dic['BTCUSD']['bid']),2)
            precos["ETH"]["compra"] = round(float(requisicao_dic['ETHUSD']['bid']),2)
            precos["ETH"]["venda"] = round(float(requisicao_dic['ETHUSD']['bid']),2)
            
            texto = f'''
            BTC {precos["BTC"]["compra"]}
            ETH {precos["ETH"]["compra"]}
            '''
            print(texto)
            infopreco.config(text=texto)
            hora = datetime.now()
            infohora.config(text=f'{hora.strftime("%H:%M:%S")}')
            if parar_sistema_event.is_set():
                break
            break   
    threading.Thread(target=tarefa, daemon=True).start()

limites = {
    "BTC": {"compra_min": 0, "venda_max": 0},
    "ETH": {"compra_min": 0, "venda_max": 0}
}
def salvar_valores():
    try:
        # Obtém os valores das entradas e atualiza os limites
        limites["BTC"]["compra_min"] = float(entry_btc_compra_min.get())
        limites["BTC"]["venda_max"] = float(entry_btc_venda_max.get())
        limites["ETH"]["compra_min"] = float(entry_eth_compra_min.get())
        limites["ETH"]["venda_max"] = float(entry_eth_venda_max.get())
        #messagebox.showinfo("Sucesso", "Limites salvos com sucesso!")
        return True

    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira valores numéricos válidos!")

    print(f"Limite Mínimo: ==compra {limites['BTC']['compra_min']}==, //compra {limites['ETH']['compra_min']}//")
    
    print(f"Limite Máximo: ==venda {limites['BTC']['venda_max']}==, //venda {limites['ETH']['venda_max']}//")
    
def iniciar_processo():
    def tarefa():
        global hora, parar_sistema
        atualizar_preco()
        time.sleep(3) 
        if salvar_valores():
            while not parar_sistema_event.is_set():
                for moeda in precos:
                    precos_venda = precos[moeda]["venda"]
                    precos_compra = precos[moeda]["compra"]

                    if precos_venda > limites[moeda]["venda_max"]:
                        atualizar_crono()
                        print(f'\033[1;31m=== ALERTA DE VENDA===\033[0m')
                        print(f'\033[1;31m{moeda}: Preço atual ({precos_venda}) ultrapassou o limite máximo ({limites[moeda]["venda_max"]} as {hora.strftime("%H:%M:%S")}).\033[0m')

                        hora = datetime.now()

                        time.sleep(10)
                        try:
                            workbook = openpyxl.load_workbook('contatos-notificados.xlsx')
                            pagina_contatos = workbook['Plan1']

                            for linha in pagina_contatos.iter_rows(min_row=5):
                                #nome, telefone
                                nome = linha[2].value    
                                telefone = linha[3].value
                                if nome and telefone:
                                #mensagem que será enviada aos contatos
                                    msg = f'Olá {nome}, hora de vender {moeda}, o limite máximo para venda foi ultrapassado({limites[moeda]["venda_max"]}) às {hora.strftime("%d/%m/%Y, %H:%M")}'

                                    #link que abrirá o zap
                                    link_mensagem_zap = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(msg)}'

                                    webbrowser.open(link_mensagem_zap)
                                    time.sleep(20)

                                    pyautogui.press('enter')
                                    time.sleep(5)
                                    print(f'Mensagem enviada para Nome= {nome} Cel= {telefone} com sucesso.')
                                    pyautogui.hotkey('ctrl', 'w')
                                    time.sleep(5)
                                else:
                                    print(f'Contato inválido Nome= {nome} Cel= {telefone}')

                        except Exception as e:
                            print(f'Erro em enviar mensagem de venda para {nome}')
        
                    elif precos_compra < limites[moeda]["compra_min"]:
                        atualizar_crono()
                        print(f'\033[1;31m=== ALERTA DE COMPRA===\033[0m')
                        print(f'\033[1;31m{moeda}: Preço atual ({precos_compra}) ultrapassou o limite mínimo ({limites[moeda]["compra_min"]} as {hora.strftime("%H:%M:%S")}).\033[0m')
                        hora = datetime.now()

                        #webbrowser.open('https://web.whatsapp.com/')
                        time.sleep(10)
                        try:
                            workbook = openpyxl.load_workbook('contatos-notificados.xlsx')
                            pagina_contatos = workbook['Plan1']

                            for linha in pagina_contatos.iter_rows(min_row=5):
                                #nome, telefone
                                nome = linha[2].value    
                                telefone = linha[3].value
                                if nome and telefone:
                                #mensagem que será enviada aos contatos
                                    msg = f'Olá {nome}, hora de comprar {moeda}, o limite mínimo para compra foi ultrapassado({limites[moeda]["compra_min"]}) às {hora.strftime("%d/%m/%Y, %H:%M")}'

                                    #link que abrirá o zap
                                    link_mensagem_zap = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(msg)}'

                                    webbrowser.open(link_mensagem_zap)
                                    time.sleep(20)

                                    pyautogui.press('enter')
                                    time.sleep(5)
                                    print(f'Mensagem enviada para Nome= {nome} Cel= {telefone} com sucesso.')
                                    pyautogui.hotkey('ctrl', 'w')
                                    time.sleep(5)
                                else:
                                    print(f'Contato inválido! Nome= {nome} Cel= {telefone}')

                        except Exception as e:
                            print(f'Erro em enviar mensagem de compra para {nome}')

                    else:
                        print('Sistema estável')
                atualizar_crono()
                janela.after(3 * 60 * 1000,iniciar_processo)
                break               

    threading.Thread(target=tarefa, daemon=True).start()

def atualizar_crono():

    if start_time:
        tempo_passado = datetime.now() - start_time
        minutos, segundos = divmod(tempo_passado.seconds,60)
        #atualiza o tempo no cronometro
        texto2 = f'Sistema iniciado há {minutos}:{segundos:02d} min'
        crono_label.config(text=texto2)
        # Reagenda a próxima atualização em 1 segundo
        janela.after(1000, atualizar_crono)

def iniciar_sistema():
    global start_time
    if not start_time:
        start_time = datetime.now()
        print('SISTEMA REINICIADO!')
        iniciar_processo()
        atualizar_crono()

def iniciar_manter_sistema():
    global thread_manter_sistema

    # Verifica se já existe uma thread ativa
    if thread_manter_sistema and thread_manter_sistema.is_alive():
        print("Manter sistema já está em execução!")
        return

    def tarefa():
        global parar_sistema_event
        while not parar_sistema_event.is_set():
            for tempo in range(30, 0, -10):
                print(f'Faltam {tempo} segundos para reiniciar o sistema')
                time.sleep(10)
            print('SISTEMA REINICIADO!')
            iniciar_sistema()
            break  # Sai do loop após reiniciar o sistema

    # Cria uma nova thread para manter_sistema, se não existir
    thread_manter_sistema = threading.Thread(target=tarefa, daemon=True)
    thread_manter_sistema.start()

def fechar_sistema():

    parar_sistema_event.set()#Sinaliza para encerrar o sistema
    for thread in threading.enumerate():
        if thread is not threading.main_thread():
            thread.join(timeout=1)
    print('FECHANDO SISTEMA')
    janela.destroy()
    time.sleep(3)
    sys.exit()   

#JANELA ABERTA
janela = Tk()#janela aberta
janela.geometry("450x555")
janela.config(bg='#161616')#cor de fundo da janela
janela.title('Alertas compra e venda de criptos  v2.0')
janela.iconbitmap('icone-sistema.ico')

# Usando pack na janela principal
top_frame = Frame(janela, bg='#161616')
top_frame.pack()

info = Label(top_frame,
    text='Favor abrir Whatsapp Web antes de iniciar o programa',
    fg='#d7ad01',
    bg='#161616',
    font=("Bebas Neue", 12)
    ).pack(pady=5)

hora = datetime.now()

infohora = Label(top_frame,
    text=f'{hora.strftime('%H:%M:%S')}',
    fg='white',
    bg='#161616',
    font=("Helvetica", 12,
    "underline"))
infohora.pack(pady=5)

infopreco = Label(top_frame,
    text="Clique no botão para atualizar e validar os preços",
    fg='white',
    bg='#161616',
    font=("Helvetica", 12))
infopreco.pack(pady=5)

atualiza_preco = Button(janela,
    text='Visualizar cotações atuais',
    fg='white',
    bg='#161616',
    highlightbackground='#d7ad01',
    font=('Times', 10),
    command=atualizar_preco)
atualiza_preco.pack(pady=5)

#3 definir valor baixo para comprar e valor alto pra vender
comando =Label(janela,
    text='Digite os valores pedidos abaixo no formato (105000.00, 3850.43)',
    fg='white',
    bg='#161616',
    font=("Cambria", 11))
comando.pack(pady=5)

# Usando grid dentro do frame
grid_frame = Frame(janela, bg='#161616')
grid_frame.pack()

#label e entry btc

Label(grid_frame, text="BTC - Compra:",
    fg='white',
    bg='#161616',  
    font=("Cambria", 11)).grid(row=0, column=0, padx=5, pady=5)
    
entry_btc_compra_min = Entry(grid_frame,
    bg='white',)

entry_btc_compra_min.grid(row=0, column=1, padx=5, pady=5)

#label e entry btc
Label(grid_frame, text="BTC - Venda :",
    fg='white',
    bg='#161616', 
    font=("Cambria", 11)).grid(row=1, column=0, padx=5, pady=5)
entry_btc_venda_max = Entry(grid_frame)
entry_btc_venda_max.grid(row=1, column=1, padx=5, pady=5)

#label e entry eth
Label(grid_frame, text="ETH - Compra:",
    fg='white',
    bg='#161616', 
    font=("Cambria", 11)).grid(row=2, column=0, padx=5, pady=5)
entry_eth_compra_min = Entry(grid_frame)
entry_eth_compra_min.grid(row=2, column=1, padx=5, pady=5)

#label e entry eth
Label(grid_frame, text="ETH - Venda :",
    fg='white',
    bg='#161616', 
    font=("Cambria", 11)).grid(row=3, column=0, padx=5, pady=5)
entry_eth_venda_max = Entry(grid_frame)
entry_eth_venda_max.grid(row=3, column=1, padx=5, pady=5)

mid_frame = Frame(janela, bg='#161616')
mid_frame.pack()
info3 =Label(mid_frame,
    text='Clique em OK para continuar o sistema',
    fg='white',
    bg='#161616', 
    font=("Cambria", 11)).pack(pady = 5)

#botão de confirmação
iniciar = Button(mid_frame,
    text='OK',
    fg='white',
    bg='#161616',
    command = iniciar_sistema).pack(pady = 5)

info4= Label(mid_frame,
    text='Caso queira parar o programa clique no botão abaixo',
    fg='white',
    bg='#161616', 
    font=("Cambria", 11)).pack(pady = 5)

parar = Button(mid_frame,
    text='Parar',
    fg='white',
    bg='#161616',
    command = fechar_sistema).pack(pady = 5)

crono_label = Label(mid_frame,
    text=f'Sistema não iniciado',
    fg='#d7ad01',
    bg='#161616',
    font=("Bebas Neue", 10))
crono_label.pack(pady=5, side='left')

bottom_frame = Frame(janela, bg='#161616')
bottom_frame.pack()
rodape = Label(bottom_frame,
    text=f'                             © Copyright 2025|Todos os Direitos Reservados a Kaích Costa                                 ',
    fg='#ffffff',
    bg='#d7ad01',
    font=("Bebas Neue", 8))
rodape.pack(pady=5)

janela.protocol("WM_DELETE_WINDOW", fechar_sistema)
janela.mainloop()