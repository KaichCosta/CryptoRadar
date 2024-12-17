
#   PASSOS
            
#5 criar um executavel pra isso tudo                               
#6 criar sistema pra enviar pro celular
#    o1 qualquer coisa criar outro arquivo pra criar um apk pro celular do meu pai e executar tudo pelo celular do meu pai e fazer ele receber as notificaçoes
#    o2 no apk, ter opção de parar a procura de preço  
#7 enviar notificação pelo email é melhor ou por zap

#1 importar bibliotecas  
import time
from datetime import datetime
import os
import yfinance as yf
import openpyxl
from urllib.parse import quote
import webbrowser
import pyautogui
from tkinter import *
import threading


#2 buscar preços bitcoin e ações atuais
symbols = ['btc-usd', 'eth-usd']
tickers = yf.Tickers(','.join(symbols))




# Variáveis globais
limite_max = []
limite_min = []
sistema_iniciado = False


#TRATAMENTO DE ERRO DECENTE
def salvar_valores():
    # Lê os valores inseridos pelo usuário
    global limite_max, limite_min
# Limpa qualquer mensagem de erro anterior
    msg_erro.config(text='')

    valores_max = entrada_max.get()
    valores_min = entrada_min.get()

    # Converte os valores para float e os armazena nas variáveis
    if ',' in valores_max and ',' in valores_min:
        try:
            valores_max = [float(valor.strip ()) for valor in valores_max.split(',')]
            valores_min = [float(valor.strip()) for valor in valores_min.split(',')]
            if len(valores_max) == 2 and len(valores_min) == 2:
                limite_max = valores_max
                limite_min = valores_min
                print(f"Limite Máximo: {limite_max}")
                print(f"Limite Mínimo: {limite_min}")
                print('VALORES SALVOS COM SUCESSO!')
                msg_erro.config(text='Valores válidos!', fg='green')
                return True  # Nenhum erro
            else:
                msg_erro.config(text='Certifique-se de fornecer exatamente dois valores, primeiro para BTC e o segundo para ETH',fg='red')
                print("Erro: Certifique-se de fornecer exatamente dois valores para cada limite.")
                return False  # Nenhum erro
        except ValueError:
            msg_erro.config(text='Digite os Preços no formato pedido',fg='red')
            print("Erro: Certifique-se de digitar os valores no formato correto (ex: 105000.00, 3850.00)")
            return False
    else:
        msg_erro.config(text='Erro: Certifique-se de separar os valores por vírgula.', fg='red') 
        return False      

def iniciar_sistema():
    if salvar_valores():
        global sistema_iniciado  # Define uma flag global para controle
        sistema_iniciado = True  # Altera o estado da flag
        ultimo_preco = [
            round(float(tickers.tickers[symbol.upper()].history(period='1d')['Close'].iloc[0]), 2)#round, float e esse ,2 são pra formatar com 2 casas decimais os valores das criptos
            for symbol in symbols
        ]
        print(f'\033[1;32m{datetime.now()}\033[0m')
        print(ultimo_preco)
        print('\033[1;32m--BITCOIN--ETHEREUM\033[0m')

        for i, preco in enumerate(ultimo_preco):
            simbolo = symbols[i].upper()

            max = limite_max[i]

            min = limite_min[i]

        if preco > max:
            print(f'\033[1;31m=== ALERTA DE VENDA===\033[0m')
            print(f'\033[1;31m{simbolo}: Preço atual ({preco}) ultrapassou o limite máximo ({max}).\033[0m')

            hora = datetime.now()

            #webbrowser.open('https://web.whatsapp.com/')
            time.sleep(10)

            workbook = openpyxl.load_workbook('contatos-notificados.xlsx')
            pagina_contatos = workbook['Plan1']
#4 CRIAR UM GRUPO E POR MEU PAI E EU PRA RECEBER ESSAS NOTIFICAÇÕES
            for linha in pagina_contatos.iter_rows(min_row=5):
                #nome, telefone
                nome = linha[2].value    
                telefone = linha[3].value

                #mensagem que será enviada aos contatos
                msg = f'Olá {nome}, hora de vender {simbolo}, {hora.strftime('%d/%m/%Y, %H:%M')}'

                #link que abrirá o zap
                link_mensagem_zap = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(msg)}'

                webbrowser.open(link_mensagem_zap)

                time.sleep(15)
                pyautogui.press('enter')
                time.sleep(5)

                pyautogui.hotkey('ctrl', 'w')
                time.sleep(2)

                pyautogui.hotkey('ctrl', 'w')
                time.sleep(5)

                print('FINALIZADO')
                break
        #==============================================
        #==============================================
        elif preco < min:
            print(f'\033[1;31m=== ALERTA DE COMPRA===\033[0m')
            print(f'\033[1;31m{simbolo}: Preço atual ({preco}) ultrapassou o limite máximo ({min}).\033[0m')
            hora = datetime.now()

            #webbrowser.open('https://web.whatsapp.com/')
            time.sleep(10)

            workbook = openpyxl.load_workbook('contatos-notificados.xlsx')
            pagina_contatos = workbook['Plan1']

            for linha in pagina_contatos.iter_rows(min_row=5):
                #nome, telefone
                nome = linha[2].value    
                telefone = linha[3].value

                #mensagem que será enviada aos contatos
                msg = f'Olá {nome}, hora de comprar {simbolo}, {hora.strftime('%d/%m/%Y, %H:%M')}'

                #link que abrirá o zap
                link_mensagem_zap = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(msg)}'

                webbrowser.open(link_mensagem_zap)

                time.sleep(15)
                pyautogui.press('enter')
                time.sleep(5)

                pyautogui.hotkey('ctrl', 'w')
                time.sleep(2)

                pyautogui.hotkey('ctrl', 'w')
                time.sleep(5)

                print('FINALIZADO')
                break
        else:
            time.sleep(300)
        threading.Thread(target=manter_sistema).start()
        
        print('SISTEMA INICIADO COM SUCESSO!')

def fechar_sistema():
    global sistema_iniciado
    sistema_iniciado = False
    print('SISTEMA FECHADO COM SUCESSO')
    janela.destroy()

def manter_sistema():
    while True:
        if sistema_iniciado: 
            print('SISTEMA ESTÁ SENDO MANTIDO COM SUCESSO!')
        else:
            print('SISTEMA DESLIGADO COM SUCESSO')
            break     
        time.sleep(10)

#JANELA ABERTA
janela = Tk()
janela.configure(bg='#f7f7f7')#cor de fundo da janela
janela.title('Preços máximos e mínimos venda e compra de criptos')

info = Label(janela, text='Favor abrir Whatsapp Web antes de iniciar o programa',fg='red',font=("Helvetica", 12, "underline")).pack(pady = 5)
#info.config(font=("Verdana", 12, "italic"))

#3 definir valor baixo para comprar e valor alto pra vender
# dar opcão do cliente escolher os valores max, min das criptos ex: limite =int(input'escolha valor min e max de eth e btc'

comando =Label(janela, text='Digite os valores pedidos abaixo no formato (105000.00, 3850.00)',font=("Helvetica", 10))
comando.pack(pady = 5)

info1 = Label(janela, text='Digite o Preço que deseja ser notificado para venda de (BTC, ETH):',font=("Helvetica", 10)).pack(pady = 1)

entrada_max = Entry(janela)
entrada_max.pack(pady=3)#caixa de digitação
# Área para exibir mensagens de erro
msg_erro = Label(janela, text='', fg='red')
msg_erro.pack(pady=1)

info2 = Label(janela, text='Digite o Preço que deseja ser notificado para compra de (BTC, ETH):',font=("Helvetica", 10)).pack(pady = 1)

entrada_min = Entry(janela)
entrada_min.pack(pady=3)#caixa de digitação

msg_erro = Label(janela, text='', fg='red')
msg_erro.pack(pady=1)

#botão de confirmação
Button(janela, text='OK',bg='#a1a1a1', command = iniciar_sistema).pack(pady = 5)

# Instrução
info3 =Label(janela, text='Clique em OK pra continuar o sistema',font=("Helvetica", 10)).pack(pady = 5)

info4= Label(janela, text='Caso queira parar o programa clique no botão abaixo',font=("Helvetica", 10)).pack(pady = 5)

Button(janela, text='Parar',bg='#a1a1a1', command = fechar_sistema).pack(pady = 5)

janela.protocol("WM_DELETE_WINDOW", fechar_sistema)

# Executar a interface gráfica em paralelo ao loop principal contido em manter sistema
janela.after(100, manter_sistema)  # Executa o loop principal depois de 100ms = 10s
janela.mainloop()


